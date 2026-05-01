#!/usr/bin/env python3
"""
validate_case.py — Aggregate LLM indicator results → pipeline grades → diff vs human grades.

Usage:
    python scripts/validate_case.py --case cases/case_01

Inputs (all inside case dir):
    adapted_indicators.xlsx   — adapted indicator set (with indicator IDs and blocks_level_up_to)
    validation/run_output.json — LLM output per indicator
    human_grades.yaml          — human assessor grades

Outputs:
    validation/diff_report.yaml — structured diff (skeleton + computed fields)

Expected run_output.json format:
    {
      "C01-L0-F2-i05": {
        "fired": true,
        "evidence_quote": "...",
        "confidence": 0.85,
        "notes": "..."
      },
      ...
    }

Aggregation logic (mirrors Aggregation_Rules sheet in master xlsx):
    For each criterion:
      1. Collect all fired indicator IDs.
      2. Check for hard blocks: any L-2 indicator with blocks_level_up_to == -2 that fired
         → caps the maximum achievable grade at -2 (hard block).
      3. Check for soft blocks: any L-1 indicator with blocks_level_up_to == -1 that fired
         → caps at -1 (cannot reach L0).
      4. Among L0 indicators: if ≥50% fired AND no hard block AND no soft block → grade = 0
      5. Among L-1 indicators: if ≥50% fired AND no hard block → grade = -1
      6. Otherwise → grade = -2
"""

import argparse
import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import yaml

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl pyyaml")


INDICATOR_ID_COL = "indicator_id"
CRITERION_COL = "criterion_id"
LEVEL_COL = "level"
BLOCKS_COL = "blocks_level_up_to"

# Column name aliases (case-insensitive, underscore-normalized)
COL_ALIASES = {
    "indicator_id": ["indicator_id", "id"],
    "criterion_id": ["criterion_id", "criterion"],
    "level": ["level"],
    "blocks_level_up_to": ["blocks_level_up_to", "blocks_level", "block"],
}

ERROR_TYPES = [
    "FN_indicator",
    "FP_indicator",
    "FN_coverage",
    "threshold_error",
    "block_error",
    "placeholder_error",
]


def normalize(s: str) -> str:
    return s.lower().strip().replace(" ", "_") if isinstance(s, str) else ""


def find_col(headers: dict, field: str) -> int | None:
    aliases = COL_ALIASES.get(field, [field])
    for alias in aliases:
        for h, idx in headers.items():
            if normalize(h) == alias:
                return idx
    return None


def load_indicators(xlsx_path: Path) -> list[dict]:
    """Load indicator rows from the Indicators sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if "indicator" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sys.exit(f"No sheet with 'indicator' in name found in {xlsx_path}")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        sys.exit("Indicators sheet is empty.")

    header_row = rows[0]
    headers = {v: i for i, v in enumerate(header_row) if v is not None}

    col_id = find_col(headers, "indicator_id")
    col_crit = find_col(headers, "criterion_id")
    col_level = find_col(headers, "level")
    col_block = find_col(headers, "blocks_level_up_to")

    missing = []
    if col_id is None:
        missing.append("indicator_id")
    if col_crit is None:
        missing.append("criterion_id")
    if col_level is None:
        missing.append("level")
    if missing:
        sys.exit(
            f"Required columns not found in Indicators sheet: {missing}\n"
            f"Available headers: {list(headers.keys())}"
        )

    indicators = []
    for row in rows[1:]:
        ind_id = row[col_id]
        if not ind_id:
            continue
        level_raw = row[col_level]
        try:
            level = int(level_raw)
        except (TypeError, ValueError):
            level = None

        block_raw = row[col_block] if col_block is not None else None
        try:
            block = int(block_raw) if block_raw is not None else None
        except (TypeError, ValueError):
            block = None

        indicators.append(
            {
                "indicator_id": str(ind_id).strip(),
                "criterion_id": str(row[col_crit]).strip() if row[col_crit] else None,
                "level": level,
                "blocks_level_up_to": block,
            }
        )

    wb.close()
    return indicators


def load_run_output(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


def load_human_grades(path: Path) -> dict:
    with open(path) as f:
        raw = yaml.safe_load(f)
    grades = {}
    for k, v in raw.items():
        if k.startswith("C") and k[1:].isdigit():
            if v == "not_assessed" or v is None:
                grades[k] = None
            else:
                try:
                    grades[k] = int(v)
                except (TypeError, ValueError):
                    grades[k] = None
    return grades


def aggregate(indicators: list[dict], run_output: dict) -> dict[str, int | None]:
    """
    Returns {criterion_id: pipeline_grade} for each criterion present in indicators.
    Grade: 0 / -1 / -2 / None (if no indicators exist for the criterion).
    """
    by_criterion: dict[str, dict] = defaultdict(
        lambda: {"L0": [], "L-1": [], "L-2": [], "blocks": []}
    )

    for ind in indicators:
        crit = ind["criterion_id"]
        if not crit:
            continue
        level = ind["level"]
        ind_id = ind["indicator_id"]
        fired = run_output.get(ind_id, {}).get("fired", False)
        block = ind["blocks_level_up_to"]

        entry = {"id": ind_id, "fired": fired, "block": block}

        if level == 0:
            by_criterion[crit]["L0"].append(entry)
        elif level == -1:
            by_criterion[crit]["L-1"].append(entry)
        elif level == -2:
            by_criterion[crit]["L-2"].append(entry)

    grades = {}
    for crit, buckets in by_criterion.items():
        l0 = buckets["L0"]
        l1 = buckets["L-1"]
        l2 = buckets["L-2"]

        # Hard block: any L-2 indicator with blocks_level_up_to == -2 that fired
        hard_blocked = any(
            e["fired"] and e["block"] == -2 for e in l2
        )

        # Soft block: any L-1 indicator with blocks_level_up_to == -1 that fired
        soft_blocked = any(
            e["fired"] and e["block"] == -1 for e in l1
        )

        # Additional soft block from L-2 indicators capping at -1
        soft_blocked = soft_blocked or any(
            e["fired"] and e["block"] == -1 for e in l2
        )

        # L0 threshold
        if l0:
            l0_fired = sum(1 for e in l0 if e["fired"])
            l0_pct = l0_fired / len(l0)
        else:
            l0_pct = 0.0

        # L-1 threshold
        if l1:
            l1_fired = sum(1 for e in l1 if e["fired"])
            l1_pct = l1_fired / len(l1)
        else:
            l1_pct = 0.0

        if l0_pct >= 0.5 and not hard_blocked and not soft_blocked:
            grades[crit] = 0
        elif l1_pct >= 0.5 and not hard_blocked:
            grades[crit] = -1
        else:
            grades[crit] = -2

    return grades


def build_diff(
    pipeline_grades: dict,
    human_grades: dict,
    criteria_in_scope: list[str],
) -> tuple[list[dict], dict]:
    mismatches = []
    matched = 0
    total = 0

    error_counts = {e: 0 for e in ERROR_TYPES}

    for crit in criteria_in_scope:
        hg = human_grades.get(crit)
        pg = pipeline_grades.get(crit)

        if hg is None:
            continue  # not assessed by human — skip

        total += 1

        if pg == hg:
            matched += 1
            continue

        delta = (pg if pg is not None else -2) - hg
        mismatches.append(
            {
                "criterion": crit,
                "human_grade": hg,
                "pipeline_grade": pg,
                "delta": delta,
                "error_types": [],  # to be filled manually
                "analysis": "",
                "fix_applied": "",
                "fix_target": "",  # master / adapted / both
            }
        )

    overall_pct = round(matched / total * 100, 1) if total else None

    summary = {
        "total_criteria_assessed": total,
        "matched": matched,
        "mismatched": len(mismatches),
        "match_pct": overall_pct,
        "mismatch_breakdown": error_counts,
        "systemic_patterns": "",
    }

    return mismatches, summary


def validate(case_dir: Path):
    case_dir = case_dir.resolve()
    val_dir = case_dir / "validation"
    val_dir.mkdir(exist_ok=True)

    # --- Load inputs ---
    adapted_xlsx = case_dir / "adapted_indicators.xlsx"
    run_output_path = val_dir / "run_output.json"
    human_grades_path = case_dir / "human_grades.yaml"
    meta_path = case_dir / "meta.yaml"

    for p in [adapted_xlsx, run_output_path, human_grades_path]:
        if not p.exists():
            sys.exit(f"Required file not found: {p}")

    print(f"Loading indicators from {adapted_xlsx.name}...")
    indicators = load_indicators(adapted_xlsx)
    print(f"  {len(indicators)} indicators loaded.")

    print(f"Loading run output from {run_output_path.name}...")
    run_output = load_run_output(run_output_path)
    print(f"  {len(run_output)} indicator results loaded.")

    print(f"Loading human grades from {human_grades_path.name}...")
    human_grades = load_human_grades(human_grades_path)

    # Load criteria_in_scope from meta.yaml
    criteria_in_scope = None
    master_version = "unknown"
    case_id = case_dir.name
    if meta_path.exists():
        with open(meta_path) as f:
            meta = yaml.safe_load(f) or {}
        criteria_in_scope = meta.get("criteria_in_scope") or []
        master_version = meta.get("master_version", "unknown")
        case_id = meta.get("case_id", case_dir.name)

    if not criteria_in_scope:
        # Fall back to all criteria mentioned in indicators
        criteria_in_scope = sorted(
            set(ind["criterion_id"] for ind in indicators if ind["criterion_id"])
        )
        print(f"Warning: criteria_in_scope not set in meta.yaml — using all {len(criteria_in_scope)} criteria from indicators.")

    # --- Coverage warning ---
    indicator_ids_in_xlsx = {ind["indicator_id"] for ind in indicators}
    result_ids = set(run_output.keys())
    missing_from_run = indicator_ids_in_xlsx - result_ids
    extra_in_run = result_ids - indicator_ids_in_xlsx
    if missing_from_run:
        print(f"\nWarning: {len(missing_from_run)} indicators in xlsx have NO result in run_output.json:")
        for i in sorted(missing_from_run)[:10]:
            print(f"  {i}")
        if len(missing_from_run) > 10:
            print(f"  ... and {len(missing_from_run) - 10} more")
    if extra_in_run:
        print(f"\nWarning: {len(extra_in_run)} results in run_output.json have NO matching indicator in xlsx:")
        for i in sorted(extra_in_run)[:5]:
            print(f"  {i}")

    # --- Aggregate ---
    print("\nAggregating indicator results to criterion grades...")
    pipeline_grades = aggregate(indicators, run_output)

    # --- Diff ---
    print("Computing diff vs human grades...")
    mismatches, summary = build_diff(pipeline_grades, human_grades, criteria_in_scope)

    # --- Print summary ---
    print(f"\n{'='*50}")
    print(f"Case: {case_id}  |  Master: {master_version}")
    print(f"Criteria assessed: {summary['total_criteria_assessed']}")
    print(f"Match: {summary['matched']} / {summary['total_criteria_assessed']} ({summary['match_pct']}%)")
    print(f"Mismatches: {summary['mismatched']}")
    if mismatches:
        print("\nMismatches:")
        for m in mismatches:
            direction = "over" if m["delta"] > 0 else "under"
            print(f"  {m['criterion']}: human={m['human_grade']}  pipeline={m['pipeline_grade']}  delta={m['delta']:+d} ({direction}estimate)")
    print(f"{'='*50}\n")

    # --- Save diff_report.yaml ---
    report = {
        "case_id": case_id,
        "master_version": master_version,
        "run_date": str(date.today()),
        "overall_match_pct": summary["match_pct"],
        "mismatches": mismatches,
        "summary": summary,
    }

    out_path = val_dir / "diff_report.yaml"
    with open(out_path, "w") as f:
        yaml.dump(report, f, allow_unicode=True, sort_keys=False, default_flow_style=False)

    print(f"Saved diff report: {out_path}")
    print("Next step: open validation/diff_report.yaml and fill in error_types, analysis, fix_applied, fix_target for each mismatch.")


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate LLM indicator results and diff vs human grades."
    )
    parser.add_argument(
        "--case",
        required=True,
        help="Path to case directory, e.g. cases/case_01",
    )
    args = parser.parse_args()
    validate(Path(args.case))


if __name__ == "__main__":
    main()
