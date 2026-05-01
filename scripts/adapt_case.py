#!/usr/bin/env python3
"""
adapt_case.py — Create a case-adapted copy of the master indicator set.

Usage:
    python scripts/adapt_case.py --case cases/case_01

What it does:
    1. Reads cases/<case_id>/placeholder_map.yaml
    2. Opens master/indicators_full_v2.xlsx
    3. Substitutes {PLACEHOLDER} → value in columns: indicator_en, adaptation_notes
    4. Warns about unfilled or unused placeholders
    5. Writes cases/<case_id>/adapted_indicators.xlsx
    6. Records current git SHA into cases/<case_id>/master_version.txt
    7. Patches cases/<case_id>/meta.yaml with master_version field
"""

import argparse
import re
import shutil
import subprocess
import sys
from pathlib import Path

import yaml

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl pyyaml")

MASTER_XLSX = Path(__file__).parent.parent / "master" / "indicators_full_v2.xlsx"
SCHEMAS_DIR = Path(__file__).parent.parent / "schemas"

# Columns in the Indicators sheet that contain placeholder text
PLACEHOLDER_COLUMNS = {"indicator_en", "adaptation_notes"}

PLACEHOLDER_PATTERN = re.compile(r"\{([A-Z_]+)\}")


def get_git_sha(repo_root: Path) -> str:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=repo_root,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip() or "unknown"
    except Exception:
        return "unknown"


def load_placeholder_map(case_dir: Path) -> dict:
    pm_path = case_dir / "placeholder_map.yaml"
    if not pm_path.exists():
        sys.exit(
            f"placeholder_map.yaml not found at {pm_path}\n"
            f"Copy from schemas/placeholder_map.yaml.template and fill in values."
        )
    with open(pm_path) as f:
        data = yaml.safe_load(f)
    return {k: (v or "") for k, v in data.items()}


def substitute(text: str, mapping: dict, used: set) -> str:
    if not isinstance(text, str):
        return text

    def replacer(m):
        key = m.group(1)
        used.add(key)
        val = mapping.get(key, "")
        if not val:
            return m.group(0)  # leave unreplaced if empty
        return val

    return PLACEHOLDER_PATTERN.sub(replacer, text)


def adapt(case_dir: Path):
    case_dir = case_dir.resolve()
    repo_root = case_dir.parent.parent

    print(f"Case dir : {case_dir}")
    print(f"Master   : {MASTER_XLSX}")

    if not case_dir.exists():
        sys.exit(f"Case directory does not exist: {case_dir}")
    if not MASTER_XLSX.exists():
        sys.exit(f"Master xlsx not found: {MASTER_XLSX}")

    placeholder_map = load_placeholder_map(case_dir)
    all_placeholders = set(placeholder_map.keys())

    # --- Substitute in xlsx ---
    wb = openpyxl.load_workbook(MASTER_XLSX)

    used_placeholders: set = set()
    unfilled_found: list = []

    indicators_sheet = None
    for sheet_name in wb.sheetnames:
        if "indicator" in sheet_name.lower():
            indicators_sheet = wb[sheet_name]
            break

    if indicators_sheet is None:
        print(
            "Warning: could not find a sheet with 'indicator' in name. "
            "Substituting in ALL sheets."
        )
        target_sheets = wb.sheetnames
    else:
        target_sheets = [indicators_sheet.title]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        # Find header row (row 1 assumed to have column names)
        headers = {
            cell.value: cell.column - 1
            for cell in ws[1]
            if cell.value
        }

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                # Check if this column should be substituted
                col_name = None
                for h, idx in headers.items():
                    if cell.column - 1 == idx and isinstance(h, str):
                        col_name = h.lower().replace(" ", "_")
                        break

                # Substitute in placeholder columns OR in any cell containing {UPPER_SNAKE}
                if col_name in PLACEHOLDER_COLUMNS or PLACEHOLDER_PATTERN.search(cell.value):
                    new_val = substitute(cell.value, placeholder_map, used_placeholders)
                    # Track unfilled
                    remaining = PLACEHOLDER_PATTERN.findall(new_val)
                    unfilled_found.extend(remaining)
                    cell.value = new_val

    # --- Save ---
    out_path = case_dir / "adapted_indicators.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # --- Git SHA ---
    sha = get_git_sha(repo_root)
    (case_dir / "master_version.txt").write_text(sha + "\n")
    print(f"Master version: {sha}")

    # --- Patch meta.yaml ---
    meta_path = case_dir / "meta.yaml"
    if meta_path.exists():
        with open(meta_path) as f:
            meta = yaml.safe_load(f) or {}
        meta["master_version"] = sha
        with open(meta_path, "w") as f:
            yaml.dump(meta, f, allow_unicode=True, sort_keys=False)
        print(f"Updated master_version in meta.yaml")

    # --- Report ---
    unused = all_placeholders - used_placeholders
    unfilled_unique = sorted(set(unfilled_found))

    print("\n--- Placeholder report ---")
    if unused:
        print(f"Unused placeholders (defined but never appeared in indicators):")
        for p in sorted(unused):
            val = placeholder_map.get(p, "")
            print(f"  {{{p}}} = '{val}'")
    else:
        print("All defined placeholders were referenced in the indicator set.")

    if unfilled_unique:
        print(f"\nUnfilled placeholders (appeared in indicators but value is empty):")
        for p in unfilled_unique:
            print(f"  {{{p}}}")
        print(
            "\nAction: fill these values in placeholder_map.yaml and re-run adapt_case.py."
        )
    else:
        print("All referenced placeholders have values. Substitution complete.")

    print("\nDone.")


def main():
    parser = argparse.ArgumentParser(description="Adapt master indicators for a specific case.")
    parser.add_argument(
        "--case",
        required=True,
        help="Path to case directory, e.g. cases/case_01",
    )
    args = parser.parse_args()
    adapt(Path(args.case))


if __name__ == "__main__":
    main()
