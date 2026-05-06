#!/usr/bin/env python3
"""
eval_case.py — Run Haiku against case transcript to produce run_output.json.

Usage:
    python scripts/eval_case.py --case cases/case_01 [--model claude-haiku-4-5] [--dry-run]

Strategy (token-efficient):
    - One API call per criterion (not per indicator) — reduces ~280 calls to ~12
    - Each call: system prompt + transcript + all indicators for that criterion
    - Haiku returns JSON: {indicator_id: {fired, evidence_quote, confidence}, ...}
    - Only sends transcript segment relevant to criterion type (skill/experience/meta)

Cost estimate (Haiku):
    - ~12 criteria × ~3K tokens input avg = ~36K tokens input total
    - ~12 criteria × ~500 tokens output avg = ~6K tokens output total
    - At Haiku pricing: negligible

Requires:
    pip install anthropic openpyxl pyyaml
    export ANTHROPIC_API_KEY=...
"""

import argparse
import json
import os
import sys
import time
from pathlib import Path

import yaml

try:
    import anthropic
    import openpyxl
except ImportError:
    sys.exit("Missing deps — run: pip install anthropic openpyxl pyyaml")


# ── Config ────────────────────────────────────────────────────────────────────

DEFAULT_MODEL = "claude-haiku-4-5"
RETRY_ATTEMPTS = 3
RETRY_DELAY = 5  # seconds between retries

SYSTEM_PROMPT = """You are an expert PM interviewer evaluating behavioral indicators.
You will receive a candidate interview transcript and a list of behavioral indicators.

For EACH indicator, determine if the behavior was OBSERVED in the transcript.

Rules:
- fired: true ONLY if there is clear, direct evidence in the transcript
- fired: false if the behavior is absent, unclear, or only implied
- evidence_quote: exact short quote from transcript (≤40 words). Empty string if fired=false.
- confidence: 0.0–1.0 (your certainty in the assessment)
- Do NOT infer. Only score what is explicitly present.
- Absence of behavior is NOT itself a behavior (unless the indicator is an absence-blocker).

Return ONLY valid JSON. No explanation, no markdown, no preamble.
Format: {"INDICATOR_ID": {"fired": bool, "evidence_quote": "...", "confidence": float}, ...}"""


# ── Load helpers ───────────────────────────────────────────────────────────────

def load_indicators(xlsx_path: Path) -> dict[str, list[dict]]:
    """Returns {criterion_id: [indicator_row, ...]}"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if "indicator" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sys.exit(f"No Indicators sheet found in {xlsx_path}")

    rows = list(sheet.iter_rows(values_only=True))
    headers = {str(v).lower().strip().replace(" ", "_"): i
               for i, v in enumerate(rows[0]) if v is not None}

    def col(name):
        aliases = {
            "indicator_id": ["indicator_id", "id"],
            "criterion_id": ["criterion_id", "criterion"],
            "level": ["level"],
            "indicator_en": ["indicator_en", "indicator"],
            "blocks_level_up_to": ["blocks_level_up_to", "blocks_level"],
        }
        for a in aliases.get(name, [name]):
            if a in headers:
                return headers[a]
        return None

    by_criterion: dict[str, list[dict]] = {}
    for row in rows[1:]:
        ind_id = row[col("indicator_id")] if col("indicator_id") is not None else None
        if not ind_id:
            continue
        crit = str(row[col("criterion_id")]).strip() if col("criterion_id") is not None else None
        level_raw = row[col("level")] if col("level") is not None else None
        try:
            level = int(level_raw)
        except (TypeError, ValueError):
            level = None
        ind_en = row[col("indicator_en")] if col("indicator_en") is not None else ""
        block_raw = row[col("blocks_level_up_to")] if col("blocks_level_up_to") is not None else None
        try:
            block = int(block_raw) if block_raw is not None else None
        except (TypeError, ValueError):
            block = None

        entry = {
            "indicator_id": str(ind_id).strip(),
            "criterion_id": crit,
            "level": level,
            "indicator_en": str(ind_en).strip() if ind_en else "",
            "blocks_level_up_to": block,
        }
        by_criterion.setdefault(crit, []).append(entry)

    wb.close()
    return by_criterion


def load_transcript(transcripts: list[str], case_dir: Path) -> str:
    parts = []
    for t in transcripts:
        p = case_dir / t
        if p.exists():
            parts.append(p.read_text(encoding="utf-8"))
        else:
            print(f"  Warning: transcript not found: {p}")
    return "\n\n---\n\n".join(parts)


def load_meta(case_dir: Path) -> dict:
    with open(case_dir / "meta.yaml") as f:
        return yaml.safe_load(f) or {}


# ── Haiku call ─────────────────────────────────────────────────────────────────

def build_user_message(criterion_id: str, criterion_name: str,
                        indicators: list[dict], transcript: str) -> str:
    ind_lines = []
    for ind in indicators:
        level_str = f"L{ind['level']}" if ind['level'] is not None else "L?"
        block_str = f" [blocks_level_up_to={ind['blocks_level_up_to']}]" if ind['blocks_level_up_to'] is not None else ""
        ind_lines.append(f"  {ind['indicator_id']} ({level_str}{block_str}): {ind['indicator_en']}")

    indicators_block = "\n".join(ind_lines)

    return f"""CRITERION: {criterion_id} — {criterion_name}

INDICATORS TO EVALUATE ({len(indicators)} total):
{indicators_block}

TRANSCRIPT:
{transcript}

Evaluate each indicator against the transcript. Return JSON only."""


def call_haiku(client, model: str, user_message: str,
               criterion_id: str, dry_run: bool) -> dict:
    if dry_run:
        print(f"    [DRY RUN] Would call {model} for {criterion_id}")
        return {}

    for attempt in range(RETRY_ATTEMPTS):
        try:
            response = client.messages.create(
                model=model,
                max_tokens=2048,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_message}],
            )
            raw = response.content[0].text.strip()

            # Strip markdown code fences if present
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            return json.loads(raw)

        except json.JSONDecodeError as e:
            print(f"    JSON parse error on attempt {attempt+1}: {e}")
            if attempt < RETRY_ATTEMPTS - 1:
                time.sleep(RETRY_DELAY)
        except Exception as e:
            print(f"    API error on attempt {attempt+1}: {e}")
            if attempt < RETRY_ATTEMPTS - 1:
                time.sleep(RETRY_DELAY)

    print(f"    FAILED after {RETRY_ATTEMPTS} attempts for {criterion_id}")
    return {}


# ── Main ───────────────────────────────────────────────────────────────────────

# Criterion name lookup (from build script)
CRITERION_NAMES = {
    "C01": "Управление бизнес моделью и монетизацией",
    "C02": "Управление бизнес моделью и монетизацией — опыт",
    "C03": "Создание и управление стратегией продукта",
    "C04": "Создание и управление стратегией продукта — опыт",
    "C05": "Управление бюджетом",
    "C06": "Управление бюджетом — опыт",
    "C07": "Планирование в горизонте года для нескольких команд",
    "C08": "Определение пользовательских метрик, построение системы метрик",
    "C09": "Определение пользовательских метрик — опыт",
    "C10": "Управление рыночными исследованиями",
    "C11": "Определение персон, потребностей и сценариев использования",
    "C12": "Проектирование пользовательского опыта",
    "C13": "Проектирование комплексных систем",
    "C14": "Проектирование комплексных систем — опыт",
    "C15": "Построение взаимоотношений и процессов коммуникаций со стейкхолдерами",
    "C16": "Приоритизация и поставка в условиях внешних дедлайнов",
    "C17": "Совместно с СТО оптимизация процесса поставки",
    "C18": "Управление продуктовым процессом в компании",
    "C19": "Создание оргструктуры и модели управления",
    "C20": "Опыт управления командой до 8 человек",
    "C21": "Находить точки роста и инсайты из исследований",
    "C22": "Работать с данными и метриками",
    "C23": "Формулировать гипотезы и описывать бизнес-требования",
    "C24": "Создание требований, сопровождение команд разработки",
    "C25": "Развитые коммуникативные навыки",
}


def evaluate(case_dir: Path, model: str, dry_run: bool):
    case_dir = case_dir.resolve()
    repo_root = case_dir.parent.parent

    adapted_xlsx = case_dir / "adapted_indicators.xlsx"
    val_dir = case_dir / "validation"
    val_dir.mkdir(exist_ok=True)
    out_path = val_dir / "run_output.json"

    meta = load_meta(case_dir)
    criteria_in_scope = meta.get("criteria_in_scope") or []
    transcripts_list = meta.get("transcripts") or []

    if not criteria_in_scope:
        sys.exit("criteria_in_scope is empty in meta.yaml")
    if not transcripts_list:
        sys.exit("transcripts list is empty in meta.yaml")

    print(f"Case     : {case_dir.name}")
    print(f"Model    : {model}")
    print(f"Criteria : {criteria_in_scope}")
    print(f"Transcripts: {transcripts_list}")

    # Load transcript
    transcript = load_transcript(transcripts_list, case_dir)
    if not transcript.strip():
        sys.exit("Transcript is empty — add files to transcripts/ folder")
    print(f"Transcript: {len(transcript)} chars")

    # Load indicators
    if not adapted_xlsx.exists():
        sys.exit(f"adapted_indicators.xlsx not found — run adapt_case.py first")
    by_criterion = load_indicators(adapted_xlsx)
    print(f"Indicators: {sum(len(v) for v in by_criterion.values())} total across {len(by_criterion)} criteria\n")

    # Load existing output (for resume/partial runs)
    existing: dict = {}
    if out_path.exists():
        with open(out_path) as f:
            existing = json.load(f)
        already_done = set()
        for ind_id in existing:
            # figure out criterion from ID prefix
            crit = "-".join(ind_id.split("-")[:1])  # e.g. C01
            already_done.add(crit)
        print(f"Resuming: {len(already_done)} criteria already in run_output.json")
    else:
        already_done = set()

    # Anthropic client
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key and not dry_run:
        sys.exit("ANTHROPIC_API_KEY not set")
    client = anthropic.Anthropic(api_key=api_key) if not dry_run else None

    results = dict(existing)

    for criterion_id in criteria_in_scope:
        if criterion_id in already_done:
            print(f"  [{criterion_id}] skipped (already evaluated)")
            continue

        indicators = by_criterion.get(criterion_id, [])
        if not indicators:
            print(f"  [{criterion_id}] no indicators found — skipping")
            continue

        criterion_name = CRITERION_NAMES.get(criterion_id, criterion_id)
        print(f"  [{criterion_id}] {criterion_name} — {len(indicators)} indicators")

        user_msg = build_user_message(criterion_id, criterion_name, indicators, transcript)

        # Token estimate
        token_estimate = len(user_msg) // 4
        print(f"    Input ~{token_estimate:,} tokens (estimate)")

        raw_result = call_haiku(client, model, user_msg, criterion_id, dry_run)

        # Normalize: ensure all indicator IDs for this criterion are present
        for ind in indicators:
            ind_id = ind["indicator_id"]
            if ind_id in raw_result:
                entry = raw_result[ind_id]
                results[ind_id] = {
                    "fired": bool(entry.get("fired", False)),
                    "evidence_quote": str(entry.get("evidence_quote", ""))[:300],
                    "confidence": float(entry.get("confidence", 0.5)),
                    "notes": str(entry.get("notes", "")),
                }
            else:
                # Model missed this indicator — default to not fired
                results[ind_id] = {
                    "fired": False,
                    "evidence_quote": "",
                    "confidence": 0.0,
                    "notes": "NOT RETURNED BY MODEL",
                }

        # Save after each criterion (resumable)
        with open(out_path, "w") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"    Done. Saved {len(results)} results so far.")

        # Brief pause between calls
        if not dry_run:
            time.sleep(1)

    print(f"\nComplete. Total results: {len(results)}")
    print(f"Saved: {out_path}")
    print(f"\nNext: python scripts/validate_case.py --case {case_dir.relative_to(repo_root)}")


def main():
    parser = argparse.ArgumentParser(description="Evaluate case transcript with Haiku.")
    parser.add_argument("--case", required=True, help="Path to case dir, e.g. cases/case_01")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"Model to use (default: {DEFAULT_MODEL})")
    parser.add_argument("--dry-run", action="store_true", help="Don't call API, just print what would happen")
    args = parser.parse_args()
    evaluate(Path(args.case), args.model, args.dry_run)


if __name__ == "__main__":
    main()
