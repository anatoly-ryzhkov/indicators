"""
Build the indicators pilot xlsx for criteria C01, C02, C25.
Output: /Users/starfish/Downloads/indicators_pilot/indicators_pilot_v1.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/starfish/Downloads/indicators_pilot/indicators_pilot_v1.xlsx"

wb = Workbook()

# ===================== STYLE HELPERS =====================
FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

L_NEG_FILL = PatternFill("solid", start_color="FCE4D6")  # -2 light red
L_BASE_FILL = PatternFill("solid", start_color="FFF2CC")  # -1 light yellow
L_STRONG_FILL = PatternFill("solid", start_color="E2EFDA")  # 0 light green

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

def style_body(ws, ncols, level_col=None):
    for row in ws.iter_rows(min_row=2, max_col=ncols):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
        if level_col is not None:
            lvl = ws.cell(row=row[0].row, column=level_col).value
            fill = None
            if lvl == -2: fill = L_NEG_FILL
            elif lvl == -1: fill = L_BASE_FILL
            elif lvl == 0: fill = L_STRONG_FILL
            if fill:
                for cell in row:
                    cell.fill = fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ===================== SHEET 1: README =====================
ws = wb.active
ws.title = "README"
readme = [
    ("Behavioral Indicators — Pilot v1", ""),
    ("", ""),
    ("Scope of pilot", "3 criteria from the source file: C01, C02, C25. Goal: validate decomposition method, indicator style, placeholder system, and aggregation logic before scaling to all 25 criteria."),
    ("", ""),
    ("Architecture", ""),
    ("Layer 1 — Indicators", "Atomic, source-agnostic, binary (fired / not fired). Stored in 'Indicators_Master'. Each indicator = one observable behavior, written to be evaluated by a fast/local LLM against transcript fragments."),
    ("Layer 2 — Orchestration (NOT in this file)", "System prompt that wraps each indicator. Handles: 'anywhere in the transcript' semantics, splitting transcript into chunks, routing skill-criteria vs experience-criteria evaluation, output schema. Designed separately."),
    ("Layer 3 — Aggregation", "Rules in 'Aggregation_Rules' sheet that combine fired/not-fired indicator results into a single level (-2 / -1 / 0) per criterion. Active anti-patterns at L-2 can block level-up via the 'blocks_level_up_to' field."),
    ("Layer 4 — Case adaptation", "Indicators contain {PLACEHOLDERS}. Before evaluation, placeholders are filled from a specific case via the 'Placeholders' sheet — each placeholder has an 'adaptation_question' that tells the adaptor (human or LLM) how to extract the right value from the case material."),
    ("", ""),
    ("Indicator anatomy", ""),
    ("Pattern", "Candidate <observable verb> <specific object> — including, but not limited to <4–6 concrete variants>, or any other <generalization>. [optional: excluding ...] [optional: e.g., '...']"),
    ("Verbs (whitelist)", "explicitly identifies / mentions / names / proposes / structures / applies / cites / asks / describes / avoids / gives / uses / connects / quantifies"),
    ("One indicator = one signal", "If you can split with AND, split. Each indicator is independently observable."),
    ("", ""),
    ("Levels", ""),
    ("-2", "Active anti-pattern observed. Evidence of failure / shallowness. May block level-up depending on 'blocks_level_up_to'."),
    ("-1", "Baseline competence behavior observed."),
    ("0", "Strong / expert behavior observed (concrete proposals, quantified forecasts, named frameworks, scaled-across-org work, mentoring, etc.)"),
    ("", ""),
    ("Decomposition logic per criterion", ""),
    ("Step 1 — Tokenize criterion text into atomic claims.", "Each comma-separated or 'и'-joined assertion is a separate claim. e.g. 'предлагал улучшение БМ И монетизации, показывал прогноз, приводил кейс' → 4 claims."),
    ("Step 2 — For each claim, generate indicators across facets.", "Facets: topic-presence / structure / evidence-type / action-proposed / tradeoff / scope / quantification."),
    ("Step 3 — Add anti-patterns derived from L-2 text.", "Active behaviors only ('avoids', 'confuses', 'gives vague answer when asked'), not absences."),
    ("Step 4 — Tag placeholders.", "Anything case-specific gets {SCREAMING_SNAKE_CASE}."),
    ("", ""),
    ("How to adapt to a new case", ""),
    ("Step 1", "Open the 'Placeholders' sheet. For each placeholder, answer the 'adaptation_question' from the case material."),
    ("Step 2", "Find/replace placeholders in the indicator_en column."),
    ("Step 3", "Re-export to LLM-prompt JSON via a script (not included here — to be built once master is approved)."),
    ("", ""),
    ("Status", "PILOT. Awaiting feedback on: granularity, indicator style, placeholder design, aggregation rules. Once approved, this template is replicated across all 25 criteria."),
]
for i, (k, v) in enumerate(readme, 1):
    ws.cell(row=i, column=1, value=k)
    ws.cell(row=i, column=2, value=v)
    if i == 1:
        ws.cell(row=i, column=1).font = Font(name=FONT, bold=True, size=14)
    elif v == "" and k != "":
        ws.cell(row=i, column=1).font = Font(name=FONT, bold=True, size=11, color="1F4E78")
    else:
        ws.cell(row=i, column=1).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=i, column=2).font = BODY_FONT
    ws.cell(row=i, column=1).alignment = WRAP
    ws.cell(row=i, column=2).alignment = WRAP
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 110

# ===================== SHEET 2: CRITERIA_SOURCE =====================
ws = wb.create_sheet("Criteria_Source")
hdrs = ["criterion_id", "criterion_name_ru", "kind", "level_-2_text", "level_-1_text", "level_0_text"]
ws.append(hdrs)
criteria = [
    ("C01", "Управление бизнес моделью и монетизацией", "skill",
     "Не использовал финансовые метрики при анализе и не смог привести убедимый пример роста прибыли/выручки в своей зоне ответственности.",
     "Корректно оперировал финансовыми показателями и связывал решения с примерами из своей практики.",
     "Предлагал улучшения бизнес-модели и монетизации, показывал прогноз эффекта и приводил детальный кейс роста прибыли/выручки."),
    ("C02", "Управление бизнес моделью и монетизацией — опыт", "experience",
     "На прямой вопрос об опыте роста прибыли ответил размыто или отрицательно.",
     "Сопоставляет принятые решения с примерами из своей практики.",
     "Детально рассказывает кейс в котором вырастил прибыль сервиса за который был ответственен."),
    ("C25", "Развитые коммуникативные навыки", "meta-skill",
     "Отвечал запутанно; интервьюеру приходилось «тянуть» дискуссию и прояснять мысли.",
     "Реактивно отвечал прозрачно и по делу, корректно и понятно закрывал вопросы по мере поступления.",
     "Вел дискуссию проактивно: обозначал план ответа, структурировал мысль и уверенно отвечал на сложные уточнения."),
]
for c in criteria:
    ws.append(c)
style_header(ws, len(hdrs))
style_body(ws, len(hdrs))
set_widths(ws, [12, 38, 14, 50, 50, 60])
for r in range(2, 2 + len(criteria)):
    ws.row_dimensions[r].height = 90

# ===================== SHEET 3: INDICATORS_MASTER =====================
ws = wb.create_sheet("Indicators_Master")
hdrs = [
    "indicator_id", "criterion_id", "level", "facet", "atomic_claim",
    "indicator_en", "placeholders_used", "blocks_level_up_to",
    "adaptation_notes", "example_positive_quote", "example_negative_quote"
]
ws.append(hdrs)

# ---- Indicators (compact list-of-tuples) ----
# Tuple: (id, crit, level, facet, atomic_claim, text, placeholders, blocks_up_to, adapt_notes, ex_pos, ex_neg)

I = []  # indicators list

# ============ C01 — skill: business model & monetization ============
I += [
    ("C01-L-2-F1-i01", "C01", -2, "financial-metrics-absence",
     "не использовал финансовые метрики при анализе",
     "Candidate discusses product impact, prioritization, or proposed changes without referencing any financial metric — including, but not limited to revenue, margin, ARPU, LTV, CAC, payback period, P&L lines, or any other monetary outcome from {FINANCIAL_METRICS_DOMAIN} — relying instead on engagement, usage, or qualitative arguments alone.",
     "{FINANCIAL_METRICS_DOMAIN}", -2,
     "{FINANCIAL_METRICS_DOMAIN} = list the 5–10 financial metrics most relevant to the case domain; use this same list across all C01/C02 indicators.",
     "—",
     "'We need to fix the dashboard latency because users are frustrated.' (no $ link)"),

    ("C01-L-2-F2-i02", "C01", -2, "example-absence",
     "не смог привести убедимый пример роста прибыли/выручки",
     "When directly asked about a case of revenue or profit growth in their area of responsibility, candidate gives a vague, deflecting, or non-answer — including, but not limited to switching to team-level achievements without isolating their contribution, naming only non-financial outcomes, saying they don't recall specifics, or any other response that fails to deliver a concrete profit/revenue growth example.",
     "—", -1,
     "Triggers only after a direct question. Orchestration layer must mark the question turn.",
     "—",
     "'We worked a lot on revenue topics, lots of things moved.'"),

    ("C01-L-2-F3-i03", "C01", -2, "financial-fluency-failure",
     "корректное оперирование финансовыми показателями (нарушено)",
     "Candidate confuses or misuses core financial concepts — including, but not limited to conflating revenue with profit, margin with markup, GMV with net revenue, treating gross numbers as net, or any other category mistake that signals shallow financial fluency.",
     "—", -2,
     "Hard block: this is a fluency floor.",
     "—",
     "'Our margin grew 30% — revenue went from 10 to 13.'"),

    ("C01-L-1-F1-i04", "C01", -1, "financial-metric-usage",
     "корректно оперировал финансовыми показателями",
     "Candidate correctly references at least one domain-relevant financial metric when reasoning about a problem or decision — including, but not limited to {FINANCIAL_METRICS_DOMAIN} — and uses it consistently with its standard definition.",
     "{FINANCIAL_METRICS_DOMAIN}", None,
     "Same placeholder as i01. 'Correctly' = no fluency-failure (cf. i03).",
     "'Take rate is 12%, so paid-services revenue is roughly GMV × 0.12.'",
     "—"),

    ("C01-L-1-F1-i05", "C01", -1, "action-to-financial-link",
     "связывал решения с финансовым исходом",
     "Candidate connects a proposed action or decision to its expected financial outcome — including, but not limited to stating that fixing X will recover revenue, that reducing Y will improve margin, that prioritizing Z protects ARPU, or any other explicit cause-to-financial-outcome link — without yet quantifying the size of the effect.",
     "—", None,
     "L-1 version: directional link only. Quantified version is L0 (i11).",
     "'Stabilizing the tax report will reduce churn of paid sellers, protecting paid-services revenue.'",
     "—"),

    ("C01-L-1-F2-i06", "C01", -1, "experience-anchor-light",
     "связывал решения с примерами из своей практики",
     "Candidate anchors a recommendation in a brief reference to their own past practice — including, but not limited to 'we did this at {OWN_PRODUCT}', 'in my last role we tried X', a one-line analogy to a prior situation, or any other short experiential anchor — without yet providing a full structured case.",
     "{OWN_PRODUCT}", None,
     "{OWN_PRODUCT} is a name-of-anything the candidate previously owned; from interview opener.",
     "'At my previous fintech we hit the same churn pattern after a price test.'",
     "—"),

    ("C01-L-1-F3-i07", "C01", -1, "financial-artifact-literacy",
     "корректное чтение P&L / unit economics",
     "Candidate reads or interprets a financial artifact in the case material correctly — including, but not limited to identifying which line drives a trend, distinguishing fixed vs variable costs, separating one-off vs recurring revenue, calling out gross vs net, or any other accurate parsing of a P&L, unit-economics table, or revenue breakdown.",
     "—", None,
     "Only fires if such an artifact exists in the case. If case has no P&L, this indicator is N/A.",
     "'The drop in this row is variable cost moving with volume — not a structural change.'",
     "—"),

    ("C01-L0-F1-i08", "C01", 0, "business-model-improvement-proposal",
     "предлагал улучшения бизнес-модели",
     "Candidate proposes a concrete change to {PRODUCT}'s business model — including, but not limited to redefining customer segments, restructuring revenue streams, changing the role of a key partner, shifting cost structure, repositioning the value proposition for a new payer, or any other component-level change from {BUSINESS_MODEL_COMPONENTS} — and the proposal is specific enough to act on, not a generic 'we should rethink the model'.",
     "{PRODUCT}, {BUSINESS_MODEL_COMPONENTS}", None,
     "{BUSINESS_MODEL_COMPONENTS} = enumerate which BM components the case actually exposes (e.g. partner bank, paid services tier, seller fees).",
     "'I'd separate large sellers into a B2B tier with bundled services and a flat fee.'",
     "—"),

    ("C01-L0-F1-i09", "C01", 0, "monetization-improvement-proposal",
     "предлагал улучшения монетизации",
     "Candidate proposes a concrete change to {PRODUCT}'s monetization mechanism — including, but not limited to introducing or removing a paid tier, repricing, bundling, switching between subscription/transactional/ad-based models, gating a free feature, or any other mechanism-level change from {MONETIZATION_LEVERS} — and the proposal is specific enough to test.",
     "{PRODUCT}, {MONETIZATION_LEVERS}", None,
     "{MONETIZATION_LEVERS} = list monetization mechanisms in the case domain.",
     "'Move tax report into a paid add-on at $5/seller/month; keep core dashboard free.'",
     "—"),

    ("C01-L0-F2-i10", "C01", 0, "forecast-presence",
     "показывал прогноз эффекта",
     "Candidate presents a forecast of the financial effect of a proposed change — including, but not limited to estimated revenue uplift, expected margin shift, projected impact on ARPU or LTV, payback timeline, or any other forward-looking quantitative claim about the proposal's outcome.",
     "—", None,
     "Forecast can be a range; this indicator only checks presence of the forward claim.",
     "'This should recover roughly $0.5–1M ARR within two quarters.'",
     "—"),

    ("C01-L0-F2-i11", "C01", 0, "forecast-quantification",
     "прогноз количественный",
     "Candidate's forecast is quantified with at least one number, range, or percentage — including, but not limited to '+5–8% revenue', 'recovers ~$1M annualized', 'breakeven in 2 quarters', or any other numeric estimate — rather than a directional 'it should help' claim.",
     "—", None,
     "Pair-fires with i10. If i10 fires but i11 doesn't, candidate has direction but no magnitude.",
     "'Conservatively, +7% paid conversion within one quarter.'",
     "'It should improve revenue noticeably.'"),

    ("C01-L0-F2-i12", "C01", 0, "forecast-assumptions",
     "обоснованность прогноза",
     "Candidate states the assumption(s) behind their forecast or names sensitivity drivers — including, but not limited to 'assuming 30% adoption', 'if churn stays flat', 'this depends on price elasticity', or any other explicit assumption or driver that conditions the projected effect.",
     "—", None,
     "Tests forecast rigor. Independent of i11 — a directional forecast can still have stated assumptions.",
     "'+7% — assumes adoption of at least 25% among current paid users.'",
     "—"),

    ("C01-L0-F3-i13", "C01", 0, "own-case-structured",
     "детальный кейс роста прибыли (структура)",
     "Candidate provides a detailed past case of revenue or profit growth in their own area of responsibility — including, but not limited to a specific product/feature/initiative they owned, a defined time window, a starting and ending state, and a problem they were solving — going beyond a one-line anchor.",
     "—", None,
     "Differs from C01-L-1-i06 (light anchor) by completeness of structure.",
     "—",
     "—"),

    ("C01-L0-F3-i14", "C01", 0, "own-case-personal-contribution",
     "личный вклад в кейсе",
     "In a past case of revenue or profit growth, candidate explicitly isolates their personal contribution — including, but not limited to 'I proposed and led', 'I owned the pricing decision', 'I personally negotiated with finance', or any other clear separation of their actions from team-wide credit.",
     "—", None,
     "—", "—", "—"),

    ("C01-L0-F3-i15", "C01", 0, "own-case-causal-mechanism",
     "механизм роста объяснён",
     "In a past case of revenue or profit growth, candidate explains the causal mechanism — including, but not limited to which lever they pulled, why the metric moved as a result, what user behavior changed, or any other explicit chain from action to financial outcome — rather than reporting only inputs and outputs.",
     "—", None,
     "—", "—", "—"),

    ("C01-L0-F3-i16", "C01", 0, "own-case-quantified-outcome",
     "результат кейса измерим",
     "In a past case of revenue or profit growth, candidate cites a measurable outcome with numbers — including, but not limited to absolute revenue lift, percentage margin improvement, multiple of baseline, or any other quantified result — not just 'it grew significantly'.",
     "—", None, "—", "—", "—"),

    ("C01-L0-F4-i17", "C01", 0, "tradeoff-articulation",
     "трейдоффы предложения",
     "Candidate references trade-offs or risks of the proposed monetization or business-model change — including, but not limited to risk of cannibalization, expected churn from a price change, brand impact of a new paywall, regulatory exposure, or any other downside acknowledged alongside the upside.",
     "—", None,
     "Captures the 'трейдоффы / развилки' signal that distinguishes a strategic L0 from a feature-level L-1.",
     "'Paywalling tax reports risks ~3–5% churn among small sellers — net still positive but worth monitoring.'",
     "—"),
]

# ============ C02 — experience: business model & monetization ============
I += [
    ("C02-L-2-F1-i01", "C02", -2, "vague-experience-answer",
     "размыто отвечал на прямой вопрос об опыте роста прибыли",
     "When directly asked about their experience growing profit or revenue, candidate gives a vague or evasive answer — including, but not limited to 'we worked on revenue topics', 'profit grew over time', 'I contributed to commercial outcomes', deferring to team-wide narratives, or any other reply that lacks a specific service, time window, or outcome.",
     "—", -1,
     "Orchestration layer must surface the direct-question turn.",
     "—",
     "'Yeah, revenue was always part of the conversation in our team.'"),

    ("C02-L-2-F1-i02", "C02", -2, "negative-experience-answer",
     "отрицательно отвечал на прямой вопрос об опыте роста прибыли",
     "When directly asked about profit-growth experience, candidate explicitly states they don't have such experience or cannot name a single instance — including, but not limited to 'I haven't owned revenue', 'profit wasn't my responsibility', 'I can't think of an example right now', or any other denial or null response.",
     "—", 0,
     "Hard block to L0. L-1 may still be reachable via other indicators if the candidate later recovers.",
     "—",
     "'Honestly, profit wasn't really in my scope.'"),

    ("C02-L-1-F1-i03", "C02", -1, "decision-to-financial-anchor",
     "сопоставляет решения с примерами из своей практики",
     "Candidate brings up at least one past situation where their decision led to a financial outcome — including, but not limited to a pricing decision they shaped, a feature they prioritized for revenue, a cost cut they pushed for, a deal they unblocked, or any other action with a stated financial consequence — even if briefly described.",
     "—", None, "—", "—", "—"),

    ("C02-L-1-F1-i04", "C02", -1, "owned-product-named",
     "явно назван свой продукт и зона ответственности",
     "Candidate names {OWN_PRODUCT} they were responsible for and identifies their role in its commercial outcomes — including, but not limited to specifying scope ownership, which P&L line they influenced, which segment they shaped, or any other concrete framing of their commercial mandate — without yet providing a structured case.",
     "{OWN_PRODUCT}", None, "—", "—", "—"),

    ("C02-L0-F1-i05", "C02", 0, "structured-profit-case",
     "структурированный кейс роста прибыли",
     "Candidate tells a structured profit-growth case for a service they owned — including, but not limited to specifying the service, the starting financial state, the trigger or problem, the chosen interventions, and the resulting financial state — covering all of these elements rather than only some.",
     "—", None,
     "Strong version requires all five elements. If only some — fires C02-L-1-i03/i04 instead.",
     "—", "—"),

    ("C02-L0-F1-i06", "C02", 0, "before-after-quantified",
     "числовой before/after",
     "In the profit-growth case, candidate quantifies both the starting and ending state — including, but not limited to baseline revenue/margin/ARPU and the post-intervention number, percentage uplift over a stated period, multiple of starting metric, or any other before-and-after numeric comparison.",
     "—", None, "—", "—", "—"),

    ("C02-L0-F1-i07", "C02", 0, "causal-logic-in-experience-case",
     "причинно-следственная логика в кейсе",
     "In the profit-growth case, candidate explains the causal logic linking interventions to the financial outcome — including, but not limited to which user behavior shifted, which monetization lever responded, why the chosen sequence mattered, or any other explicit mechanism — rather than juxtaposing actions and results without connection.",
     "—", None, "—", "—", "—"),

    ("C02-L0-F1-i08", "C02", 0, "personal-accountability-in-experience-case",
     "личная ответственность в кейсе",
     "Candidate explicitly isolates their personal accountability in the profit-growth case — including, but not limited to 'I owned the P&L for this service', 'I personally proposed the new pricing', 'I led the cross-team initiative', or any other clear separation of their authority and contribution from collective work.",
     "—", None, "—", "—", "—"),

    ("C02-L0-F1-i09", "C02", 0, "course-correction-honesty",
     "признаки реального опыта (что не получилось)",
     "Candidate names what didn't work or what they had to course-correct in the profit-growth case — including, but not limited to a failed first attempt, a hypothesis that was wrong, a metric that initially moved the wrong way, a stakeholder pushback, or any other concrete reflection on the path — signaling a real lived case rather than a polished narrative.",
     "—", None,
     "Strong authenticity signal; rare in fabricated cases.",
     "—", "—"),

    ("C02-L0-F1-i10", "C02", 0, "scale-and-horizon-context",
     "масштаб и горизонт кейса",
     "Candidate cites the time horizon and scale context of the profit-growth case — including, but not limited to over how many quarters or years, against what team size or budget, in what market or business stage, or any other framing that lets the listener calibrate the magnitude of the achievement.",
     "—", None, "—", "—", "—"),
]

# ============ C25 — meta-skill: communication ============
I += [
    ("C25-L-2-F1-i01", "C25", -2, "avoids-direct-answer",
     "отвечает не на тот вопрос",
     "Candidate avoids directly answering the question and instead delivers loosely related background or stories — including, but not limited to long company-context narratives before reaching the point, anecdotes that don't address the asked question, deflecting to adjacent topics, or any other response that fails to deliver an actual answer.",
     "—", -1, "—", "—",
     "Q: 'How did you choose this priority?' — A: 'Well, our company has a long history with this kind of problem...'"),

    ("C25-L-2-F1-i02", "C25", -2, "no-logical-structure",
     "ответ без логической структуры",
     "Candidate gives responses that lack logical structure, making it hard to follow the main point — including, but not limited to jumping between unrelated facts, mixing examples with arguments without separation, leaving sentences unfinished, or any other delivery without clear framing or progression.",
     "—", -1, "—", "—", "—"),

    ("C25-L-2-F1-i03", "C25", -2, "irrelevant-detail-overload",
     "лишние детали размывают мысль",
     "Candidate includes irrelevant or excessive detail that dilutes the main message — including, but not limited to going off-topic mid-answer, piling technical details before establishing the core idea, listing context that doesn't serve the question, or any other overload that obscures the answer.",
     "—", -1, "—", "—", "—"),

    ("C25-L-2-F2-i04", "C25", -2, "vague-language",
     "размытый язык",
     "Candidate uses vague, imprecise, or placeholder language that makes ideas unclear — including, but not limited to frequent 'things', 'stuff', 'that thing we did', 'some kind of', extensive 'um/uh' fillers, or any other low-content language in place of specific terms.",
     "—", -1, "—", "—",
     "'We did some stuff with the data thing and it kind of worked.'"),

    ("C25-L-2-F1-i05", "C25", -2, "untracked-topic-jumps",
     "перескакивает между темами",
     "Candidate jumps between topics without signaling transitions or finishing previous points — including, but not limited to starting one explanation then mid-sentence shifting to another, abandoning a thread without summarizing, weaving multiple unresolved threads, or any other untracked topic switching.",
     "—", -1, "—", "—", "—"),

    ("C25-L-2-F3-i06", "C25", -2, "interviewer-pulls-answer",
     "интервьюер вытягивает ответ",
     "Interviewer has to repeatedly pull or extract the answer through follow-ups — including, but not limited to asking the same question two or more times to get a usable response, restating the question to clarify it, prompting 'so what is your answer?', or any other pattern where the interviewer carries the conversational load.",
     "—", -1,
     "Requires the orchestration layer to expose interviewer turns separately from candidate turns.",
     "—", "—"),

    ("C25-L-2-F4-i07", "C25", -2, "no-pause-no-checkin",
     "не делает пауз, не проверяет понимание",
     "Candidate does not pause or check in during long monologues — including, but not limited to speaking for two or more minutes without inviting questions, ignoring visible cues to wrap up, missing the interviewer's attempts to interject, or any other failure to manage shared airtime.",
     "—", 0,
     "Blocks L0 (proactive facilitation), but not L-1.",
     "—", "—"),

    ("C25-L-1-F1-i08", "C25", -1, "direct-answer",
     "отвечает прямо и по делу",
     "Candidate answers questions directly and explicitly, without unnecessary background or storytelling — including, but not limited to leading with the answer ('Yes, we did X', 'The reason is Y'), closing the question before adding context, or any other on-topic delivery.",
     "—", None, "—",
     "Q: 'Why this priority?' — A: 'Because it has the largest revenue impact. Specifically, ...'", "—"),

    ("C25-L-1-F1-i09", "C25", -1, "reactive-clarity",
     "ясность по мере поступления вопросов",
     "Candidate's answers are clear and on-point as questions arrive — including, but not limited to addressing each question discretely, not chaining multiple unrelated answers into one block, leaving the interviewer satisfied without re-asking, or any other reactive clarity.",
     "—", None, "—", "—", "—"),

    ("C25-L-1-F2-i10", "C25", -1, "precise-terms",
     "использует точные термины",
     "Candidate uses precise terms instead of placeholder language — including, but not limited to naming the actual metric, system, role, or framework rather than 'the thing', 'a tool', 'someone from another team', or any other concrete substitution for vague reference.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F1-i11", "C25", 0, "states-answer-plan",
     "обозначает план ответа",
     "Candidate proactively states a plan for their answer before delivering it — including, but not limited to 'Let me break this down into three parts', 'I'll start with X, then move to Y', 'first the diagnosis, then the proposal', or any other explicit upfront roadmap of the response.",
     "—", None, "—",
     "'Let me cover three things: what I'd diagnose first, what I'd propose, and the trade-offs.'", "—"),

    ("C25-L0-F1-i12", "C25", 0, "uses-signposts",
     "использует сигнпосты и заголовки",
     "Candidate uses signposts or headlines to mark important points and guide the listener — including, but not limited to 'the key takeaway is', 'three reasons why', 'the most important thing here is', 'to summarize', or any other explicit emphasis or structure marker.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F1-i13", "C25", 0, "intro-transitions-summary",
     "вступление, переходы, итог",
     "Candidate structures responses with a clear opening, transitions, and a closing summary — including, but not limited to 'let me set context, then propose, then mention risks', 'to wrap up', 'so in short', or any other end-to-end framing applied within a single answer.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F2-i14", "C25", 0, "simplifies-complex-ideas",
     "объясняет сложное просто",
     "Candidate explains complex ideas using simple language, analogies, or concrete examples — including, but not limited to walking through a step-by-step example, using an analogy to a familiar domain, replacing jargon with plain terms when it helps clarity, or any other simplification of an abstract idea.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F3-i15", "C25", 0, "checks-clarity",
     "проактивно проверяет понимание",
     "Candidate proactively checks if their answer was clear or offers to expand — including, but not limited to 'does that answer your question?', 'want me to go deeper on any of those?', 'should I expand on the second point?', or any other explicit comprehension check.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F3-i16", "C25", 0, "paces-with-pauses",
     "паузы и приглашение вопросов",
     "Candidate pauses or invites questions during long answers to give space for clarification — including, but not limited to natural pauses after each segment, 'any questions on this part before I move on?', stepwise checking in after a major point, or any other intentional pacing of shared airtime.",
     "—", None, "—", "—", "—"),

    ("C25-L0-F4-i17", "C25", 0, "handles-pushback-confidently",
     "уверенно отвечает на сложные уточнения",
     "Candidate handles difficult clarifying or pushback questions confidently — including, but not limited to acknowledging the challenge before answering, restating the underlying question to confirm understanding, defending a position with reasoning rather than retreating, or any other composed response to scrutiny.",
     "—", None, "—",
     "'That's a fair challenge. Let me restate the assumption — I'm assuming X. Given that, I still think Y because...'", "—"),
]

for row in I:
    ws.append(row)

style_header(ws, len(hdrs))
style_body(ws, len(hdrs), level_col=3)
set_widths(ws, [18, 12, 8, 32, 38, 95, 30, 18, 50, 55, 55])
for r in range(2, 2 + len(I)):
    ws.row_dimensions[r].height = 110

# ===================== SHEET 4: PLACEHOLDERS =====================
ws = wb.create_sheet("Placeholders")
hdrs = ["placeholder", "type", "description", "adaptation_question", "example_value_for_marketplace_case"]
ws.append(hdrs)
placeholders = [
    ("{PRODUCT}", "scalar (string)",
     "The product or service that is the subject of the case.",
     "What product/service is the candidate analyzing in this case? Use the product's actual name as it appears in the case material.",
     "SortItOut Today"),
    ("{OWN_PRODUCT}", "scalar (string)",
     "The candidate's actual past product/service brought up during the experience interview.",
     "Filled at evaluation time, not adaptation time. The orchestration layer extracts this from the candidate's own intro/answers; if absent, the indicator does not fire.",
     "(extracted live)"),
    ("{FINANCIAL_METRICS_DOMAIN}", "list (5–10 items)",
     "The financial metrics that are most relevant in the case domain. Used as anchors in C01-L-2-i01 and C01-L-1-i04.",
     "Reading the case material, list the 5–10 financial metrics most relevant here (e.g., for a marketplace: GMV, take rate, paid-services revenue, ARPU, LTV, CAC, contribution margin, gross profit). Pick metrics actually visible or implied by the case data.",
     "GMV, take rate, paid-services revenue, ARPU, LTV, contribution margin, EBITDA, payback period"),
    ("{MONETIZATION_LEVERS}", "list (3–7 items)",
     "Monetization mechanisms that exist or could be introduced in the case domain. Used in C01-L0-i09.",
     "What monetization mechanisms does the case product use or could plausibly use? List the levers.",
     "transaction fees, paid services subscription, premium tier, paywalled add-ons, ads, partner referrals"),
    ("{BUSINESS_MODEL_COMPONENTS}", "list (3–7 items)",
     "Business-model components actually exposed by the case (per Business Model Canvas style decomposition). Used in C01-L0-i08.",
     "Which BM components does the case actually expose to candidate's reasoning? (customer segments, key partners, revenue streams, cost structure, value proposition, channels)",
     "seller segments, partner bank, paid-services revenue stream, infrastructure cost structure, B2B value proposition"),
]
for p in placeholders:
    ws.append(p)
style_header(ws, len(hdrs))
style_body(ws, len(hdrs))
set_widths(ws, [32, 20, 50, 60, 50])
for r in range(2, 2 + len(placeholders)):
    ws.row_dimensions[r].height = 90

# ===================== SHEET 5: AGGREGATION_RULES =====================
ws = wb.create_sheet("Aggregation_Rules")
hdrs = ["rule_id", "criterion_id", "rule_text", "rationale"]
ws.append(hdrs)
rules = [
    ("R-GENERAL-1", "ALL",
     "Indicators are independent binary observations. Aggregation produces a single level per criterion: -2, -1, 0, or 'insufficient_evidence'.",
     "Keeps indicator authoring decoupled from leveling logic."),
    ("R-GENERAL-2", "ALL",
     "Hard block: if a fired L-2 indicator has blocks_level_up_to=-2, the criterion's level is capped at -2 regardless of other fires.",
     "Some negative behaviors (concept confusion, never using financial metrics) make higher levels meaningless."),
    ("R-GENERAL-3", "ALL",
     "Soft block: if a fired L-2 indicator has blocks_level_up_to=-1, the criterion cannot be assigned 0 (but L-1 remains reachable).",
     "Some L-2 patterns (e.g. one vague answer) are recoverable if compensated by L-1 evidence elsewhere."),
    ("R-GENERAL-4", "ALL",
     "L0 assignment: at least 50% of the criterion's L0 indicators fire AND no hard or soft L-2 block is active.",
     "50% threshold tunable. Pilot default."),
    ("R-GENERAL-5", "ALL",
     "L-1 assignment: at least 50% of the criterion's L-1 indicators fire AND L0 condition not met AND no hard L-2 block.",
     "Same threshold pattern."),
    ("R-GENERAL-6", "ALL",
     "L-2 assignment: any L-2 indicator fires AND L-1/L0 conditions not met.",
     "—"),
    ("R-GENERAL-7", "ALL",
     "Insufficient evidence: none of the above conditions met (e.g., the topic was never raised in the relevant transcript pool).",
     "Distinguishes 'we didn't observe' from 'we observed failure'."),
    ("R-C01-1", "C01",
     "C01 indicators are evaluated against the case-discussion transcript pool only.",
     "Skill criterion — assesses reasoning shown in the case."),
    ("R-C02-1", "C02",
     "C02 indicators are evaluated against the experience-interview transcript pool only.",
     "Experience criterion — assesses prior-doing claims."),
    ("R-C25-1", "C25",
     "C25 indicators are evaluated against the entire transcript (case + experience).",
     "Communication is a meta-skill present throughout the session."),
    ("R-C25-2", "C25",
     "C25-L-2-F3-i06 (interviewer-pulls-answer) requires turn-level speaker labels in the transcript; if turns are unlabeled, this indicator is N/A.",
     "Indicator depends on who said what; not all transcripts carry that signal."),
    ("R-PAIR-1", "C01+C02",
     "Optional cross-criterion composition: if C02 = 0 (rich experience case), this can be cited as supporting evidence when judging C01 borderline cases. Not automatic — handled by the human reviewer or a separate composition prompt.",
     "Per user direction: criteria work independently but may compose at the leveling stage."),
]
for r in rules:
    ws.append(r)
style_header(ws, len(hdrs))
style_body(ws, len(hdrs))
set_widths(ws, [16, 14, 70, 55])
for r in range(2, 2 + len(rules)):
    ws.row_dimensions[r].height = 80

# ===================== SAVE =====================
wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Total indicators: {len(I)}")
print(f"  C01: {len([x for x in I if x[1]=='C01'])}")
print(f"  C02: {len([x for x in I if x[1]=='C02'])}")
print(f"  C25: {len([x for x in I if x[1]=='C25'])}")
