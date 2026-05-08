"""
build_corp_view.py — Regroup PM indicators by corporate competency clusters.

Reads master/indicators_full_v2.xlsx, remaps all 284 PM indicators to 6 corporate
competency clusters at the INDICATOR level (not criterion level), appends 9 new
AI-literacy indicators, writes corp_view_v1.xlsx.

Output: master/corp_view_v1.xlsx
  Sheet "CorpView"  — all 293 rows sorted by corp_competency → level → indicator_id
  Sheet "Summary"   — count per competency + criteria represented
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR        = os.path.dirname(__file__)
MASTER_XLSX = os.path.join(DIR, "indicators_full_v2.xlsx")
OUT_XLSX    = os.path.join(DIR, "corp_view_v1.xlsx")

# ── Competency codes (short) ────────────────────────────────────────────────────
КЦ  = "Клиентоцентричность"
РКС = "Развитие команд и сотрудничество"
УС  = "Управление собой"
УРО = "Управление результатом и ответственность"
СМ  = "Системное мышление и решение проблем"
AI  = "AI грамотность"

CORP_ORDER = [КЦ, РКС, УС, УРО, СМ, AI]

# ── Indicator-level mapping: indicator_id → corp_competency ───────────────────
# Every existing indicator assigned based on its text, not its parent criterion.
IND_MAP = {
    # ── C01 Business Model & Monetization ────────────────────────────────────
    "C01-L-2-F1-i01": УРО,  # no financial metrics in reasoning
    "C01-L-2-F2-i02": УРО,  # vague on revenue/profit growth
    "C01-L-2-F3-i03": СМ,   # confuses financial concepts (analytical flaw)
    "C01-L-1-F1-i04": УРО,  # references financial metric
    "C01-L-1-F1-i05": УРО,  # connects action to financial outcome
    "C01-L-1-F2-i06": УРО,  # anchors in past practice
    "C01-L-1-F3-i07": СМ,   # reads/interprets financial artifact
    "C01-L0-F1-i08":  УРО,  # proposes business model change
    "C01-L0-F1-i09":  УРО,  # proposes monetization mechanism change
    "C01-L0-F2-i10":  СМ,   # financial forecast
    "C01-L0-F2-i11":  СМ,   # quantified forecast
    "C01-L0-F2-i12":  СМ,   # states assumptions behind forecast
    "C01-L0-F3-i13":  УРО,  # past revenue/profit growth case
    "C01-L0-F3-i14":  УРО,  # isolates personal contribution
    "C01-L0-F3-i15":  СМ,   # explains causal mechanism
    "C01-L0-F3-i16":  УРО,  # cites measurable outcome
    "C01-L0-F4-i17":  СМ,   # trade-offs/risks of monetization change

    # ── C02 Business Model Experience ────────────────────────────────────────
    "C02-L-2-F1-i01": УРО,  # vague on revenue-growth experience
    "C02-L-2-F1-i02": УРО,  # explicitly no revenue experience
    "C02-L-1-F1-i03": УРО,  # past situation with financial outcome
    "C02-L-1-F1-i04": УРО,  # names own product, role in commercial outcomes
    "C02-L0-F1-i05":  УРО,  # structured profit-growth case
    "C02-L0-F1-i06":  УРО,  # quantifies start and end state
    "C02-L0-F1-i07":  СМ,   # explains causal logic
    "C02-L0-F1-i08":  УРО,  # isolates personal accountability
    "C02-L0-F1-i09":  УС,   # names what didn't work (self-reflection, learning)
    "C02-L0-F1-i10":  УРО,  # cites time horizon and scale

    # ── C03 Product Strategy ─────────────────────────────────────────────────
    "C03-L-2-F1-i01": УРО,  # cannot articulate value proposition
    "C03-L-2-F1-i02": УРО,  # cannot articulate strategy above feature level
    "C03-L-1-F1-i03": УРО,  # articulates value proposition
    "C03-L-1-F1-i04": КЦ,   # ties VP to specific target audience + need
    "C03-L-1-F2-i05": УРО,  # states product direction beyond features
    "C03-L0-F1-i06":  УРО,  # unique value proposition vs alternatives
    "C03-L0-F2-i07":  УРО,  # states what is sacrificed/de-prioritized
    "C03-L0-F2-i08":  СМ,   # explains mechanism of strategic gain
    "C03-L0-F2-i09":  СМ,   # explains why now (market window reasoning)
    "C03-L0-F3-i10":  СМ,   # alternative strategic forks considered
    "C03-L0-F4-i11":  КЦ,   # maps in-scope vs out-of-scope segments
    "C03-L0-F5-i12":  УРО,  # connects strategy to business outcome
    "C03-L0-F5-i13":  СМ,   # references named strategy framework

    # ── C04 Strategy Experience ───────────────────────────────────────────────
    "C04-L-2-F1-i01": УРО,  # vague on own product strategy
    "C04-L-1-F1-i02": УРО,  # names strategic goal
    "C04-L-1-F1-i03": КЦ,   # names target audience of strategy
    "C04-L-1-F1-i04": УРО,  # states value proposition
    "C04-L-1-F1-i05": УРО,  # names key strategic directions
    "C04-L-1-F1-i06": УРО,  # names period priorities
    "C04-L0-F1-i07":  УРО,  # initiated/rebuilt strategic vision
    "C04-L0-F2-i08":  СМ,   # justifies with data
    "C04-L0-F2-i09":  СМ,   # justifies via explicit trade-offs
    "C04-L0-F3-i10":  РКС,  # drove alignment of strategy with stakeholders
    "C04-L0-F4-i11":  УРО,  # traces strategy to plan changes
    "C04-L0-F4-i12":  УРО,  # cites measurable results

    # ── C05 Budget Management ─────────────────────────────────────────────────
    "C05-L-2-F1-i01": УРО,  # no costs referenced in proposals
    "C05-L-2-F1-i02": УРО,  # non-answer on cost structure
    "C05-L-1-F1-i03": СМ,   # references cost analytics (analytical)
    "C05-L-1-F1-i04": СМ,   # explains cost line item drivers
    "C05-L-1-F1-i05": СМ,   # distinguishes cost types
    "C05-L0-F1-i06":  УРО,  # proposes concrete cost optimization
    "C05-L0-F2-i07":  УРО,  # quantifies financial effect
    "C05-L0-F3-i08":  СМ,   # identifies optimization risks
    "C05-L0-F3-i09":  СМ,   # growth-hurting vs neutral optimization
    "C05-L0-F3-i10":  СМ,   # frames as trade-off

    # ── C06 Budget Experience ─────────────────────────────────────────────────
    "C06-L-2-F1-i01": УРО,  # cannot answer on cost structure of own product
    "C06-L-1-F1-i02": СМ,   # knows unit economics
    "C06-L-1-F1-i03": СМ,   # references market cost benchmarks
    "C06-L-1-F1-i04": СМ,   # describes cost structure of own product
    "C06-L0-F1-i05":  УРО,  # cites budget optimization case
    "C06-L0-F1-i06":  УРО,  # cites measurable outcome
    "C06-L0-F1-i07":  УРО,  # isolates personal contribution
    "C06-L0-F1-i08":  СМ,   # names trade-offs encountered in optimization

    # ── C07 Annual Planning ───────────────────────────────────────────────────
    "C07-L-2-F1-i01": УРО,  # fails to deliver a plan
    "C07-L-2-F1-i02": УРО,  # plan ignores team capacity
    "C07-L-2-F1-i03": УРО,  # plan is unrealistic
    "C07-L-1-F1-i04": УРО,  # short-term plan with sequenced activities
    "C07-L-1-F2-i05": СМ,   # decision points/forks in plan
    "C07-L-1-F3-i06": УРО,  # grounded in available resources
    "C07-L-1-F3-i07": РКС,  # grounded in engineering/team estimates (working with team)
    "C07-L0-F1-i08":  УРО,  # long-horizon plan (≥6 months)
    "C07-L0-F2-i09":  СМ,   # states explicit planning approach
    "C07-L0-F3-i10":  СМ,   # risk management in plan
    "C07-L0-F4-i11":  СМ,   # maps dependencies explicitly
    "C07-L0-F5-i12":  РКС,  # plan covers and coordinates multiple teams
    "C07-L0-F6-i13":  СМ,   # named planning framework

    # ── C08 User Metrics ──────────────────────────────────────────────────────
    "C08-L-2-F1-i01": СМ,   # only satisfaction proxies (analytical flaw)
    "C08-L-2-F2-i02": СМ,   # cannot link metric to business outcomes
    "C08-L-1-F1-i03": КЦ,   # identifies key behavioral user metric
    "C08-L-1-F1-i04": КЦ,   # explains what user behavior the metric captures
    "C08-L-1-F1-i05": СМ,   # justifies metric for this product specifically
    "C08-L0-F1-i06":  СМ,   # proposes metric system/hierarchy
    "C08-L0-F2-i07":  СМ,   # shows causal chain metric → business outcome
    "C08-L0-F3-i08":  СМ,   # named metric framework
    "C08-L0-F4-i09":  СМ,   # leading vs lagging indicators
    "C08-L0-F5-i10":  УРО,  # ties user metric to revenue/margin
    "C08-L0-F6-i11":  СМ,   # identifies missing metrics
    "C08-L-2-F3-i12": СМ,   # flat metric collection without hierarchy

    # ── C09 User Metrics Experience ───────────────────────────────────────────
    "C09-L-2-F1-i01": СМ,   # only satisfaction proxies (experience)
    "C09-L-2-F1-i02": СМ,   # cannot describe metric system in own product
    "C09-L-1-F1-i03": КЦ,   # names key user metric of own product
    "C09-L-1-F1-i04": СМ,   # states formula/calculation method
    "C09-L-1-F1-i05": КЦ,   # explains user behavior the metric reflects
    "C09-L0-F1-i06":  СМ,   # multi-layer metric system in own product
    "C09-L0-F2-i07":  СМ,   # relationships between metrics
    "C09-L0-F3-i08":  УРО,  # connects user-metric system to business metrics
    "C09-L0-F4-i09":  УРО,  # cites decision driven by these metrics
    "C09-L0-F4-i10":  СМ,   # case of measuring effect of a change

    # ── C10 Market Research ───────────────────────────────────────────────────
    "C10-L-2-F1-i01": КЦ,   # claims without research (ignores customer signals)
    "C10-L-2-F1-i02": КЦ,   # dismisses market research as low-value
    "C10-L-1-F1-i03": КЦ,   # references market research findings
    "C10-L-1-F1-i04": КЦ,   # cites specific research source
    "C10-L-1-F1-i05": КЦ,   # applies research findings to decisions
    "C10-L0-F1-i06":  СМ,   # critically evaluates research methodology
    "C10-L0-F2-i07":  СМ,   # identifies gaps/errors in research
    "C10-L0-F3-i08":  КЦ,   # proposes additional research sources
    "C10-L0-F4-i09":  СМ,   # distinguishes correlation from causation
    "C10-L0-F5-i10":  СМ,   # cross-references multiple sources
    "C10-L0-F6-i11":  СМ,   # proposes better interpretation of research

    # ── C11 Personas, Needs, Scenarios ────────────────────────────────────────
    "C11-L-2-F1-i01": КЦ,   # claims without user grounding
    "C11-L-2-F1-i02": КЦ,   # cannot articulate persona
    "C11-L-1-F1-i03": КЦ,   # references persona
    "C11-L-1-F2-i04": КЦ,   # ties decision to user need
    "C11-L-1-F1-i05": КЦ,   # distinguishes two distinct personas
    "C11-L0-F1-i06":  КЦ,   # ties decisions to specific user scenarios
    "C11-L0-F2-i07":  КЦ,   # identifies gaps in persona definitions
    "C11-L0-F3-i08":  КЦ,   # proposes persona/need/scenario corrections
    "C11-L0-F4-i09":  КЦ,   # cites primary research for persona claims
    "C11-L0-F5-i10":  КЦ,   # maps scenarios across user lifecycle

    # ── C12 UX Design ─────────────────────────────────────────────────────────
    "C12-L-2-F1-i01": КЦ,   # cannot describe interface
    "C12-L-2-F1-i02": КЦ,   # no UX dimension at all
    "C12-L-1-F1-i03": КЦ,   # feature list without flow
    "C12-L-1-F2-i04": КЦ,   # tasks without flows
    "C12-L0-F1-i05":  КЦ,   # coherent layout
    "C12-L0-F2-i06":  КЦ,   # user scenario step by step
    "C12-L0-F3-i07":  КЦ,   # UX choices tied to value proposition
    "C12-L0-F4-i08":  КЦ,   # UX choices tied to metrics
    "C12-L0-F5-i09":  КЦ,   # considers entry points and CTAs
    "C12-L0-F6-i10":  КЦ,   # addresses error/edge cases in UX

    # ── C13 Complex Systems ───────────────────────────────────────────────────
    "C13-L-2-F1-i01": СМ,   # purely additive without system integrity
    "C13-L-2-F2-i02": СМ,   # ignores downstream impact
    "C13-L-1-F1-i03": СМ,   # improving existing vs adding new
    "C13-L-1-F2-i04": СМ,   # weak prioritization of problems
    "C13-L0-F1-i05":  СМ,   # explicit prioritization among problems
    "C13-L0-F2-i06":  СМ,   # proposes simplification
    "C13-L0-F2-i07":  СМ,   # removes/sunsets component
    "C13-L0-F3-i08":  СМ,   # maps user/system roles in architecture
    "C13-L0-F3-i09":  СМ,   # maps dependencies across components
    "C13-L0-F4-i10":  СМ,   # trade-offs of architectural choices
    "C13-L0-F5-i11":  СМ,   # architecture in terms of user scenarios

    # ── C14 Complex Systems Experience ───────────────────────────────────────
    "C14-L-2-F1-i01": СМ,   # no complex system case
    "C14-L-2-F1-i02": СМ,   # case lacks complexity
    "C14-L-1-F1-i03": СМ,   # case with multiple roles/components/dependencies
    "C14-L-1-F2-i04": УС,   # acknowledges past solution was not fully coherent (self-honesty)
    "C14-L-1-F3-i05": СМ,   # proposes path to system coherence
    "C14-L0-F1-i06":  СМ,   # achieved coherence
    "C14-L0-F2-i07":  СМ,   # explains key trade-offs in case
    "C14-L0-F3-i08":  СМ,   # defines system boundaries
    "C14-L0-F4-i09":  СМ,   # component coherence mechanisms
    "C14-L0-F4-i10":  РКС,  # stakeholder alignment on architecture
    "C14-L0-F5-i11":  СМ,   # regulatory/compliance dimension
    "C14-L0-F5-i12":  СМ,   # vendor/integration dimension

    # ── C15 Stakeholder Communication ────────────────────────────────────────
    "C15-L-2-F1-i01": РКС,  # no stakeholder consideration
    "C15-L-2-F2-i02": РКС,  # cannot describe conflict resolution
    "C15-L-1-F1-i03": РКС,  # acknowledges stakeholders at resource level
    "C15-L-1-F1-i04": РКС,  # lists stakeholders affected
    "C15-L0-F1-i05":  РКС,  # addresses stakeholders at people level
    "C15-L0-F1-i06":  РКС,  # addresses at process level
    "C15-L0-F1-i07":  РКС,  # addresses at goals level
    "C15-L0-F2-i08":  РКС,  # regular communication rituals
    "C15-L0-F3-i09":  УС,   # addresses stakeholder conflict explicitly (requires EQ, courage)
    "C15-L0-F4-i10":  РКС,  # manages expectations proactively
    "C15-L0-F5-i11":  РКС,  # maps stakeholders by influence/interest

    # ── C16 Prioritization under deadlines ───────────────────────────────────
    "C16-L-2-F1-i01": УРО,  # plan ignores team estimates
    "C16-L-2-F1-i02": УРО,  # plan ignores constraints
    "C16-L-1-F1-i03": РКС,  # plan based on team estimates (works with team)
    "C16-L-1-F2-i04": СМ,   # accounts for dependencies
    "C16-L-1-F3-i05": УРО,  # no contingency for slippage
    "C16-L0-F1-i06":  СМ,   # accounts for risks in prioritization
    "C16-L0-F2-i07":  УРО,  # proposes scope fork
    "C16-L0-F2-i08":  СМ,   # proposes sequencing fork
    "C16-L0-F2-i09":  РКС,  # proposes resource fork (coordinates resources)
    "C16-L0-F3-i10":  РКС,  # proposes team productivity improvements
    "C16-L0-F4-i11":  УРО,  # addresses slip scenarios explicitly

    # ── C17 Delivery Process with CTO ────────────────────────────────────────
    "C17-L-2-F1-i01": СМ,   # no delivery process awareness
    "C17-L-2-F2-i02": СМ,   # cannot name delivery bottlenecks
    "C17-L-1-F1-i03": СМ,   # describes delivery process in detail
    "C17-L-1-F1-i04": СМ,   # identifies specific bottlenecks
    "C17-L-1-F1-i05": СМ,   # knows release cadence and gates
    "C17-L0-F1-i06":  УРО,  # cites delivery improvement case (result)
    "C17-L0-F2-i07":  УРО,  # states personal contribution
    "C17-L0-F3-i08":  СМ,   # cites delivery metrics
    "C17-L0-F4-i09":  УРО,  # cites measurable result of improvement
    "C17-L0-F5-i10":  СМ,   # named delivery framework

    # ── C18 Product Process ───────────────────────────────────────────────────
    "C18-L-2-F1-i01": УРО,  # cannot describe product decision-making
    "C18-L-2-F2-i02": КЦ,   # product decisions without research (ignores users)
    "C18-L-1-F1-i03": СМ,   # references industry product practices
    "C18-L-1-F2-i04": СМ,   # applies frameworks superficially
    "C18-L-1-F2-i05": СМ,   # doesn't evaluate framework fit
    "C18-L0-F1-i06":  СМ,   # identifies errors in product artifacts
    "C18-L0-F2-i07":  СМ,   # states what data/artifact to trust vs not
    "C18-L0-F3-i08":  РКС,  # proposes concrete process improvement (team-level)
    "C18-L0-F4-i09":  СМ,   # context-specific framework choice with rationale
    "C18-L0-F5-i10":  РКС,  # corrects colleagues' misapplication (coaching)

    # ── C19 Org Structure ─────────────────────────────────────────────────────
    "C19-L-2-F1-i01": РКС,  # no hard managerial decision case
    "C19-L-2-F1-i02": УС,   # deflects on firing/restructuring (avoids difficulty)
    "C19-L-1-F1-i03": РКС,  # systematic team growth
    "C19-L-1-F2-i04": РКС,  # retention work
    "C19-L-1-F3-i05": РКС,  # defines roles
    "C19-L-1-F3-i06": РКС,  # assigns accountability
    "C19-L0-F1-i07":  РКС,  # multiple-X team growth
    "C19-L0-F2-i08":  РКС,  # management model (role + scope + KPI + cadence)
    "C19-L0-F3-i09":  РКС,  # scalable management model
    "C19-L0-F4-i10":  УС,   # difficult separation handled with care (EQ under pressure)
    "C19-L0-F4-i11":  РКС,  # reorg case

    # ── C20 Team Management ───────────────────────────────────────────────────
    "C20-L-2-F1-i01": РКС,  # no direct management example
    "C20-L-2-F1-i02": РКС,  # no people responsibility in prior role
    "C20-L-1-F1-i03": РКС,  # direct management of 1–3 people
    "C20-L-1-F2-i04": РКС,  # goal-setting with direct reports
    "C20-L-1-F3-i05": РКС,  # regular 1:1s
    "C20-L-1-F4-i06": РКС,  # giving feedback
    "C20-L0-F1-i07":  РКС,  # direct management of 4–8 people
    "C20-L0-F2-i08":  РКС,  # hiring and development
    "C20-L0-F3-i09":  РКС,  # distributing accountability
    "C20-L0-F4-i10":  РКС,  # resolving conflict in team
    "C20-L0-F5-i11":  РКС,  # systemic management approach

    # ── C21 Growth Insights from Research ────────────────────────────────────
    "C21-L-2-F1-i01": КЦ,   # research without changing direction (doesn't act on insights)
    "C21-L-2-F1-i02": УС,   # dismisses conflicting insights (cognitive rigidity)
    "C21-L-1-F1-i03": КЦ,   # identifies problem point from research
    "C21-L-1-F2-i04": КЦ,   # names problem but no growth opportunity
    "C21-L0-F1-i05":  КЦ,   # concrete growth opportunity from research
    "C21-L0-F2-i06":  УРО,  # opportunity in actionable, testable form
    "C21-L0-F3-i07":  КЦ,   # tied to specific user segment
    "C21-L0-F4-i08":  УРО,  # sizes the growth opportunity
    "C21-L0-F5-i09":  УРО,  # translates insight to strategy/plan change

    # ── C22 Data & Metrics ────────────────────────────────────────────────────
    "C22-L-2-F1-i01": СМ,   # claims without data
    "C22-L-2-F1-i02": СМ,   # ignores data in case material
    "C22-L-1-F1-i03": СМ,   # decision based on metric
    "C22-L-1-F1-i04": СМ,   # explains how to interpret metric
    "C22-L-1-F1-i05": СМ,   # leading vs lagging distinction
    "C22-L0-F1-i06":  СМ,   # identifies selection/cohort bias
    "C22-L0-F1-i07":  СМ,   # accounts for seasonality
    "C22-L0-F2-i08":  СМ,   # questions metric definition
    "C22-L0-F3-i09":  СМ,   # identifies measurement error
    "C22-L0-F3-i10":  СМ,   # proposes fix for measurement
    "C22-L0-F4-i11":  СМ,   # triangulates multiple metrics
    "C22-L0-F5-i12":  СМ,   # correlation vs causation

    # ── C23 Hypotheses & Requirements ────────────────────────────────────────
    "C23-L-2-F1-i01": СМ,   # no hypothesis framing
    "C23-L-2-F2-i02": СМ,   # cannot articulate what is being tested
    "C23-L-1-F1-i03": СМ,   # formulates hypothesis
    "C23-L-1-F2-i04": УРО,  # states basic business requirements
    "C23-L-1-F3-i05": СМ,   # no validation plan
    "C23-L-1-F3-i06": УРО,  # no success criteria
    "C23-L0-F1-i07":  СМ,   # concrete validation method
    "C23-L0-F2-i08":  УРО,  # numeric success criteria
    "C23-L0-F3-i09":  УРО,  # prioritizes hypotheses by effect/effort
    "C23-L0-F4-i10":  СМ,   # states falsification condition
    "C23-L0-F5-i11":  УРО,  # links hypothesis to business outcome

    # ── C24 Requirements & Dev Support ───────────────────────────────────────
    "C24-L-2-F1-i01": УРО,  # requirements can't let team start
    "C24-L-2-F2-i02": УРО,  # no acceptance criteria
    "C24-L-2-F3-i03": УРО,  # no priorities in requirements
    "C24-L-1-F1-i04": УРО,  # clear, unambiguous requirements
    "C24-L-1-F2-i05": УРО,  # basic acceptance criteria
    "C24-L-1-F3-i06": СМ,   # limited risk consideration in requirements
    "C24-L-1-F4-i07": СМ,   # limited alternatives in requirements
    "C24-L0-F1-i08":  УРО,  # clear goal in requirements
    "C24-L0-F2-i09":  СМ,   # explicit scope boundaries
    "C24-L0-F3-i10":  СМ,   # names dependencies in requirements
    "C24-L0-F4-i11":  РКС,  # shadowed/supported delivery (works with dev team)
    "C24-L0-F5-i12":  РКС,  # change management during build

    # ── C25 Communication ────────────────────────────────────────────────────
    "C25-L-2-F1-i01": УС,   # avoids answering / deflects
    "C25-L-2-F1-i02": УС,   # lacks logical structure
    "C25-L-2-F1-i03": УС,   # irrelevant/excessive detail
    "C25-L-2-F2-i04": УС,   # vague placeholder language
    "C25-L-2-F1-i05": УС,   # jumps between topics
    "C25-L-2-F3-i06": УС,   # interviewer has to pull answers
    "C25-L-2-F4-i07": РКС,  # no pause/check-in during monologue (doesn't engage other)
    "C25-L-1-F1-i08": УС,   # answers directly and explicitly
    "C25-L-1-F1-i09": УС,   # answers are clear and on-point
    "C25-L-1-F2-i10": УС,   # precise terms (not placeholder language)
    "C25-L0-F1-i11":  УС,   # states plan for answer (self-organization)
    "C25-L0-F1-i12":  УС,   # uses signposts
    "C25-L0-F1-i13":  УС,   # clear opening/transitions/closing
    "C25-L0-F2-i14":  РКС,  # simple language, analogies (adapts to audience)
    "C25-L0-F3-i15":  РКС,  # checks if clear / offers to expand (engages other)
    "C25-L0-F3-i16":  РКС,  # pauses and invites questions (collaborative)
    "C25-L0-F4-i17":  УС,   # handles pushback/difficult questions confidently

    # ── C11 additions — JTBD / need translation / journey / segmentation ─────
    "C11-L-2-F2-i03": КЦ,   # solution without user need grounding
    "C11-L-1-F3-i06": КЦ,   # surface-level need interpretation (face value)
    "C11-L0-F6-i11":  КЦ,   # JTBD / translates stated wants into underlying jobs
    "C11-L0-F7-i12":  КЦ,   # multi-touchpoint / customer journey thinking
    "C11-L0-F8-i13":  КЦ,   # segmentation sophistication

    # ── C13 additions — second-order effects ─────────────────────────────────
    "C13-L-1-F3-i05": СМ,   # first-order causation only
    "C13-L0-F6-i12":  СМ,   # anticipates second-order consequences

    # ── C14 additions — growth mindset, self-awareness, accountability ────────
    "C14-L-2-F2-i03": УС,   # systematic external attribution (hard block)
    "C14-L-2-F3-i04": УС,   # self-assessment contradicts interview evidence (soft block)
    "C14-L-1-F4-i06": УС,   # surface-level failure acknowledgment without learning
    "C14-L-1-F5-i07": УС,   # reactive learning only (under task pressure)
    "C14-L0-F6-i13":  УС,   # frames failures as explicit learning events
    "C14-L0-F7-i14":  УС,   # first-person accountability for failures
    "C14-L0-F7-i15":  УС,   # proactive self-initiated learning behaviors

    # ── C15 additions — cross-functional influence without authority ──────────
    "C15-L0-F6-i12":  РКС,  # influence without authority (persuasion / coalition)

    # ── C16 additions — fail-fast discipline + customer advocacy in trade-offs
    "C16-L-2-F2-i03": УРО,  # sunk-cost continuation without stopping criteria
    "C16-L0-F5-i12":  УРО,  # predefined stopping criteria (kill-switch discipline)
    "C16-L0-F6-i13":  КЦ,   # customer advocacy in prioritization trade-offs

    # ── C18 additions — pre-mortem ───────────────────────────────────────────
    "C18-L0-F6-i11":  СМ,   # proactive pre-mortem / anticipatory failure thinking

    # ── C19 additions — productive conflict (РКС) + stress response (УС) ─────
    "C19-L-2-F2-i03": УС,   # destabilizing behavior under pressure (soft block)
    "C19-L-1-F4-i07": РКС,  # conflict avoidance / premature consensus
    "C19-L0-F5-i12":  РКС,  # productive conflict navigation to better outcome
    "C19-L0-F5-i13":  УС,   # regulated stress response (triage, reset, quality held)

    # ── C20 additions — psychological safety + capability building ────────────
    "C20-L-2-F2-i03": РКС,  # safety suppression / punishes bad news (hard block)
    "C20-L-2-F3-i04": РКС,  # people as pure execution resources (soft block)
    "C20-L-1-F5-i07": РКС,  # safety creation only in formal settings
    "C20-L0-F6-i12":  РКС,  # proactive psychological safety creation
    "C20-L0-F6-i13":  РКС,  # deliberate capability building in team members

    # ── C22 additions — goal quality (УРО) + AI metric translation (AI) ──────
    "C22-L-1-F2-i06": УРО,  # vague goal-setting without measurable criteria
    "C22-L0-F6-i13":  УРО,  # well-formed goal with target, timeframe, baseline
    "C22-L0-F7-i14":  AI,   # translates ML metrics (precision/recall) → business

    # ── C23 additions — hypothesis-first orientation ──────────────────────────
    "C23-L-2-F3-i03": СМ,   # assumption-free solution framing (soft block)
    "C23-L0-F6-i12":  СМ,   # hypothesis as starting point before solution

    # ── C26 — AI в продукте (new criterion) ──────────────────────────────────
    "C26-L-2-F1-i01": AI,   # fundamental AI misstatement (hard block)
    "C26-L-2-F2-i02": AI,   # AI bias blindness (soft block)
    "C26-L-2-F3-i03": AI,   # uncritical AI output acceptance (soft block)
    "C26-L-1-F1-i04": AI,   # AI black-box framing without product implications
    "C26-L0-F1-i05":  AI,   # calibrated AI capability mental model
    "C26-L0-F2-i06":  AI,   # responsible AI design (fairness, bias, privacy)
    "C26-L0-F3-i07":  AI,   # effective AI task/prompt design
    "C26-L0-F4-i08":  AI,   # critical evaluation of AI outputs
    "C26-L0-F5-i09":  AI,   # build/buy/partner reasoning for AI capabilities
    "C26-L0-F6-i10":  AI,   # AI team collaboration (ML engineers, data scientists)
    "C26-L0-F7-i11":  AI,   # AI factored into product strategy
    # v3.1 additions — L1–L5 framework coverage
    "C26-L-2-F4-i12": AI,   # L1: no professional AI tool usage at all (soft block)
    "C26-L-2-F5-i13": AI,   # L1: confuses AI with rule-based automation (soft block)
    "C26-L-1-F2-i05": AI,   # L2: regular simple AI usage (one tool, weekly)
    "C26-L-1-F3-i06": AI,   # L2: privacy/confidentiality risk awareness for cloud AI
    "C26-L-1-F4-i07": AI,   # L2: understands prompt / context effect on output quality
    "C26-L0-F8-i12":  AI,   # L3: workflow redesigned for AI (not AI bolted onto old flow)
    "C26-L0-F9-i13":  AI,   # L3: multi-tool practice (2–4 tools, different task types)
    "C26-L0-F9-i14":  AI,   # L3: creates custom AI configurations (GPTs / Projects)
    "C26-L0-F10-i15": AI,   # L4: deploys AI into team processes / builds automations
    "C26-L0-F10-i16": AI,   # L4: calculates and communicates ROI of AI adoption
    "C26-L0-F11-i17": AI,   # L4: teaches and enables colleagues to use AI
    "C26-L0-F12-i18": AI,   # L4: understands and applies agentic / multi-step AI workflows
    "C26-L0-F12-i19": AI,   # L4: systematic AI eval (golden sets, regression tracking)
    "C26-L0-F13-i20": AI,   # L5: AI governance / org-level change thinking
}

# AI_INDICATORS removed in v3 — AI literacy indicators now live in C26
# in the master xlsx (build_indicators_full.py) and are mapped via IND_MAP above.

# ── Style constants ────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
HDR_FILL  = PatternFill("solid", start_color="1F4E78")
HDR_FONT  = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
WRAP      = Alignment(wrap_text=True, vertical="top")
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_L_NEG    = PatternFill("solid", start_color="FCE4D6")
FILL_L_BASE   = PatternFill("solid", start_color="FFF2CC")
FILL_L_STRONG = PatternFill("solid", start_color="E2EFDA")

COMP_FILLS = {
    КЦ:  PatternFill("solid", start_color="DDEEFF"),
    РКС: PatternFill("solid", start_color="E8D5F5"),
    УС:  PatternFill("solid", start_color="D5EAF5"),
    УРО: PatternFill("solid", start_color="FDEBD0"),
    СМ:  PatternFill("solid", start_color="D5F5E3"),
    AI:  PatternFill("solid", start_color="FAD7E0"),
}


# ── Load master ────────────────────────────────────────────────────────────────
def load_master(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        if "indicator" in name.lower():
            sheet = wb[name]; break
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(v).lower().strip().replace(" ", "_") if v else "" for v in rows[0]]
    result = []
    for row in rows[1:]:
        if not row[0]: continue
        result.append(dict(zip(headers, row)))
    wb.close()
    return result


# ── Build ──────────────────────────────────────────────────────────────────────
def build():
    print(f"Loading {MASTER_XLSX} ...")
    master_rows = load_master(MASTER_XLSX)
    print(f"  {len(master_rows)} master indicators loaded.")

    # Verify coverage
    missing = [str(r["indicator_id"]).strip() for r in master_rows
               if str(r["indicator_id"]).strip() not in IND_MAP]
    if missing:
        print(f"\n  ⚠ {len(missing)} indicators NOT in IND_MAP — will be SKIPPED:")
        for m in missing:
            print(f"    {m}")
    else:
        print("  ✓ All master indicators covered in IND_MAP.")

    # Group by competency
    by_comp = {c: [] for c in CORP_ORDER}
    for row in master_rows:
        ind_id = str(row["indicator_id"]).strip()
        comp = IND_MAP.get(ind_id)
        if comp is None:
            continue
        level_raw = row["level"]
        try:
            level = int(level_raw)
        except (TypeError, ValueError):
            level = None
        block_raw = row.get("blocks_level_up_to")
        try:
            block = int(block_raw) if block_raw is not None else None
        except (TypeError, ValueError):
            block = None
        by_comp[comp].append({
            "indicator_id":    ind_id,
            "corp_competency": comp,
            "pm_criterion_id": str(row.get("criterion_id", "")).strip(),
            "level":           level,
            "facet":           row.get("facet", ""),
            "atomic_claim":    row.get("atomic_claim", ""),
            "indicator_en":    str(row.get("indicator_en", "")).strip(),
            "blocks_level_up_to": block,
        })

    # AI indicators now come from master C26 via IND_MAP — no separate append needed.

    level_order = {-2: 0, -1: 1, 0: 2}
    for comp in CORP_ORDER:
        by_comp[comp].sort(key=lambda r: (level_order.get(r["level"], 9), r["indicator_id"]))

    # ── Workbook ───────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "CorpView"

    COLS = [
        ("corp_competency",    28),
        ("indicator_id",       22),
        ("pm_criterion_id",    13),
        ("level",               7),
        ("blocks_level_up_to", 13),
        ("facet",              20),
        ("atomic_claim",       26),
        ("indicator_en",       82),
    ]

    for ci, (hdr, _) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row_num = 2
    total = 0
    comp_counts = {}

    for comp in CORP_ORDER:
        rows_for_comp = by_comp[comp]
        if not rows_for_comp:
            continue
        # Section header
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num, end_column=len(COLS))
        hcell = ws.cell(row=row_num, column=1,
                        value=f"  {comp}  ({len(rows_for_comp)} indicators)")
        hcell.fill = COMP_FILLS[comp]
        hcell.font = Font(name=FONT_NAME, bold=True, size=12)
        hcell.alignment = Alignment(vertical="center", horizontal="left")
        hcell.border = BORDER
        ws.row_dimensions[row_num].height = 22
        row_num += 1

        for ind in rows_for_comp:
            level = ind["level"]
            lf = {-2: FILL_L_NEG, -1: FILL_L_BASE, 0: FILL_L_STRONG}.get(level)
            for ci, (hdr, _) in enumerate(COLS, 1):
                cell = ws.cell(row=row_num, column=ci, value=ind.get(hdr))
                cell.font = BODY_FONT; cell.alignment = WRAP; cell.border = BORDER
                if lf: cell.fill = lf
            ws.row_dimensions[row_num].height = 60
            row_num += 1
            total += 1

        comp_counts[comp] = len(rows_for_comp)

    for ci, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Corporate Competency", "PM Criteria represented", "# Indicators"])
    COMP_CRITERIA = {
        КЦ:  "C11 (JTBD, journey, segmentation) + C16 (advocacy) + parts of C03, C04, C08, C09, C12, C18",
        РКС: "C15, C19, C20 (full) + parts of C04, C07, C14, C16, C24, C25",
        УС:  "C14 (growth mindset, accountability) + C19 (stress) + C25 (most)",
        УРО: "C01, C02 (most) + C16 (fail-fast) + C22 (goal quality) + parts of C03–C10, C17, C23, C24",
        СМ:  "C13, C22, C23 (full) + C18 (pre-mortem) + parts of C01, C03, C05–C10, C14–C18, C24",
        AI:  "C26 — AI в продукте (full) + C22-i14 (AI metric translation)",
    }
    for comp in CORP_ORDER:
        ws2.append([comp, COMP_CRITERIA.get(comp, ""), comp_counts.get(comp, 0)])
    for cell in ws2[1]:
        cell.font = Font(name=FONT_NAME, bold=True)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 14

    wb.save(OUT_XLSX)

    print(f"\nWrote {OUT_XLSX}")
    print(f"Total indicators: {total}  (328 master indicators, all mapped via IND_MAP)")
    print()
    for comp in CORP_ORDER:
        print(f"  {comp}: {comp_counts.get(comp, 0)}")


if __name__ == "__main__":
    build()
